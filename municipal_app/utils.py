import openpyxl
import pdfplumber
import re
from decimal import Decimal, InvalidOperation

def extract_total_from_excel(file):
    """
    Extrae el monto total de un archivo Excel.
    Asume que el total está en la celda A1 o busca la última fila con datos numéricos.
    """
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        # Primero intenta celda A1
        cell_value = sheet['A1'].value
        if cell_value and isinstance(cell_value, (int, float)):
            return Decimal(str(cell_value))

        # Busca la última fila con datos numéricos
        for row in range(sheet.max_row, 0, -1):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, (int, float)):
                    return Decimal(str(cell_value))

        raise ValueError("No se encontró un monto numérico en el archivo Excel.")

    except Exception as e:
        raise ValueError(f"Error procesando archivo Excel: {str(e)}")

def extract_total_from_pdf(file):
    """
    Extrae el monto total de un archivo PDF.
    Busca patrones numéricos en el texto extraído.
    """
    try:
        with pdfplumber.open(file) as pdf:
            text = ''
            for page in pdf.pages:
                text += page.extract_text() or ''

        # Busca números decimales o enteros en el texto
        # Patrón para montos: números con o sin decimales, opcionalmente con separadores
        pattern = r'(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?)'
        matches = re.findall(pattern, text)

        if matches:
            # Toma el último match como posible total
            last_match = matches[-1]
            # Limpia separadores
            cleaned = last_match.replace(',', '').replace('.', '')
            if len(cleaned) > 2:
                # Asume los últimos 2 dígitos son decimales
                cleaned = cleaned[:-2] + '.' + cleaned[-2:]
            else:
                cleaned = '0.' + cleaned.zfill(2)
            return Decimal(cleaned)

        raise ValueError("No se encontró un monto en el archivo PDF.")

    except Exception as e:
        raise ValueError(f"Error procesando archivo PDF: {str(e)}")

def extract_total_from_file(file):
    """
    Función general para extraer total de cualquier archivo soportado.
    """
    filename = file.name.lower()
    if filename.endswith(('.xlsx', '.xls')):
        return extract_total_from_excel(file)
    elif filename.endswith('.pdf'):
        return extract_total_from_pdf(file)
    else:
        raise ValueError("Tipo de archivo no soportado. Solo Excel y PDF.")